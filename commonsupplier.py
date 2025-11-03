import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill

# ============================================================
#                 HELPER FUNCTIONS
# ============================================================

def clean_headers(df):
    """Standardize column headers (remove extra spaces, non-breaking spaces)."""
    df.columns = df.columns.str.strip().str.replace("\u00A0", " ", regex=False)
    return df


def find_column(df, target_name):
    """Find column name in df matching target_name (case & space insensitive)."""
    for col in df.columns:
        if col.lower().replace(" ", "") == target_name.lower().replace(" ", ""):
            return col
    raise KeyError(f"❌ Column '{target_name}' not found in sheet.")


def update_cell(df, idx, col, new_val, modified_cells, sheet_name):
    """Update cell value and record modification if changed."""
    if col in df.columns:
        old_val = df.at[idx, col]
        if str(old_val).strip() != str(new_val).strip():
            df.at[idx, col] = new_val
            if pd.isna(old_val) and new_val == "":
                pass
            else:
                modified_cells.add((sheet_name, idx, col))


def ensure_column(df, target_col):
    """Ensure a column exists, creating it if missing."""
    if target_col not in df.columns:
        df[target_col] = ""
    return target_col


# ============================================================
#                 MAIN PROCESSING LOGIC
# ============================================================
def process_excel(input_bytes, pairs):
    input_file = BytesIO(input_bytes)
    output_file = BytesIO()

    print("🔹 Loading Excel file (header at row 2, keeping all text)...")
    sheets = pd.read_excel(input_file, sheet_name=None, header=1, dtype=str)

    required_sheets = [
        "BUT000 - General",
        "BUT100 - Role",
        "ADRC - Address",
        "LFA1 - Supplier General",
        "LFB1 - Company Code (Supplier)",
    ]
    optional_sheets = [
        "LFM1 - Purchasing Org Data",
        "WYT3 - Partner Function (Suppli",
    ]

    for s in required_sheets:
        if s not in sheets:
            raise ValueError(f"❌ Missing required sheet: {s}")

    modified_cells = set()

    # ------------------------------------------------------------
    # Step 1: Build mapping from user input
    # ------------------------------------------------------------
    id_map = {}  # child -> parent
    src_info = {}

    for parent, child in pairs:
        parent = str(parent).strip()
        child = str(child).strip()
        id_map[child] = parent
        src_info[parent] = {"type": "parent_id"}
        src_info[child] = {"type": "child_id"}

    print("✅ Parent–Child mapping successfully built.")

    # ------------------------------------------------------------
    # Step 2: Assign Role Info (same as before)
    # ------------------------------------------------------------
    but100 = clean_headers(sheets["BUT100 - Role"])
    role_src_col = find_column(but100, "Source_ID")
    src_count = but100[role_src_col].value_counts().to_dict()

    for sid, info in src_info.items():
        info["role"] = "PO" if src_count.get(sid, 0) > 1 else "NPO"

    print("✅ Source_ID attributes and roles assigned.")

    # ------------------------------------------------------------
    # Step 3: BUT000 - General
    # ------------------------------------------------------------
    print("🛠 Updating BUT000 - General...")
    but000 = clean_headers(sheets["BUT000 - General"])
    source_col = find_column(but000, "Source_ID")

    cols_clear = [
        "NAME_ORG2", "NAME_ORG3", "NAME_ORG4",
        "MC_NAME2", "MC_NAME3", "MC_NAME4",
        "ZGSTS_SLP_REP_FLG", "ZGSTS_CMT_REP_FLG", "ZGSTS_ATL_REP_FLG"
    ]
    cols_fill_x = ["ZGSTS_AVN_REP_FLG", "XDELE"]
    cols_name = ["MC_NAME1", "NAME_ORG1"]

    for idx, row in but000.iterrows():
        sid = str(row[source_col]).strip()
        if sid in src_info and src_info[sid]["type"] == "child_id":
            parent_id = id_map.get(sid, "0000000000")
            for col in cols_clear:
                if col in but000.columns:
                    update_cell(but000, idx, col, "", modified_cells, "BUT000 - General")
            for col in cols_fill_x:
                if col in but000.columns:
                    update_cell(but000, idx, col, "X", modified_cells, "BUT000 - General")
            for col in cols_name:
                update_cell(but000, idx, col, f"COMMON SUPPLIER {parent_id}", modified_cells, "BUT000 - General")
        else:
            update_cell(but000, idx, "ZGSTS_CMT_REP_FLG", "X", modified_cells, "BUT000 - General")
            update_cell(but000, idx, "ZGSTS_ATL_REP_FLG", "X", modified_cells, "BUT000 - General")

    sheets["BUT000 - General"] = but000

    # ------------------------------------------------------------
    # Step 4: ADRC - Address
    # ------------------------------------------------------------
    print("🛠 Updating ADRC - Address...")
    adrc = clean_headers(sheets["ADRC - Address"])
    adrc_src_col = find_column(adrc, "Source_ID")
    adrc_name_col = find_column(adrc, "Name1")

    for idx, row in adrc.iterrows():
        sid = str(row[adrc_src_col]).strip()
        if sid in src_info and src_info[sid]["type"] == "child_id":
            parent_id = id_map.get(sid, "0000000000")
            update_cell(adrc, idx, adrc_name_col, f"COMMON SUPPLIER {parent_id}", modified_cells, "ADRC - Address")

    sheets["ADRC - Address"] = adrc

    # ------------------------------------------------------------
    # Step 5: LFA1 - Supplier General
    # ------------------------------------------------------------
    print("🛠 Updating LFA1 - Supplier General...")
    lfa1 = clean_headers(sheets["LFA1 - Supplier General"])
    lfa1_src_col = find_column(lfa1, "Source_ID")
    cols_to_clear = ["NAME2", "NAME3", "NAME4"]
    cols_to_replace = ["NAME1"]
    cols_fill_x = ["LOEVM", "SPERR", "SPERM"]

    for idx, row in lfa1.iterrows():
        sid = str(row[lfa1_src_col]).strip()
        if sid in src_info and src_info[sid]["type"] == "child_id":
            parent_id = id_map.get(sid, "0000000000")
            for col in cols_to_clear:
                update_cell(lfa1, idx, col, "", modified_cells, "LFA1 - Supplier General")
            for col in cols_to_replace:
                update_cell(lfa1, idx, col, f"COMMON SUPPLIER {parent_id}", modified_cells, "LFA1 - Supplier General")
            for col in cols_fill_x:
                update_cell(lfa1, idx, col, "X", modified_cells, "LFA1 - Supplier General")

    sheets["LFA1 - Supplier General"] = lfa1

    # ------------------------------------------------------------
    # Step 6: LFB1 - Company Code (Supplier)
    # ------------------------------------------------------------
    print("🛠 Updating LFB1 - Company Code (Supplier)...")
    lfb1 = clean_headers(sheets["LFB1 - Company Code (Supplier)"])
    lfb1_src_col = find_column(lfb1, "Source_ID")
    bukrs_col = find_column(lfb1, "BUKRS")
    action_col = ensure_column(lfb1, "_ACTION_CODE")

    bukrs_map = lfb1.groupby(lfb1_src_col)[bukrs_col].apply(lambda x: set(x.dropna().astype(str).str.strip())).to_dict()

    for idx, row in lfb1.iterrows():
        sid = str(row[lfb1_src_col]).strip()
        if sid in src_info and src_info[sid]["type"] == "child_id":
            parent_id = id_map.get(sid, "0000000000")
            parent_bukrs = bukrs_map.get(parent_id, set())
            bukrs_val = str(row[bukrs_col]).strip()
            if bukrs_val and bukrs_val not in parent_bukrs:
                update_cell(lfb1, idx, action_col, "I", modified_cells, "LFB1 - Company Code (Supplier)")
                update_cell(lfb1, idx, lfb1_src_col, parent_id, modified_cells, "LFB1 - Company Code (Supplier)")

    sheets["LFB1 - Company Code (Supplier)"] = lfb1

    # ------------------------------------------------------------
    # Step 7: LFM1 - Purchasing Org Data
    # ------------------------------------------------------------
    sheet_name = "LFM1 - Purchasing Org Data"
    if sheet_name in sheets:
        print(f"🛠 Updating {sheet_name}...")
        df = clean_headers(sheets[sheet_name])
        try:
            src_col = find_column(df, "Source_ID")
            ekorg_col = find_column(df, "EKORG")
        except KeyError as e:
            print(f"⚠️ {e}")
            src_col = ekorg_col = None

        if src_col and ekorg_col:
            action_col = ensure_column(df, "_ACTION_CODE")
            ekorg_map = df.groupby(src_col)[ekorg_col].apply(lambda x: set(x.dropna().astype(str).str.strip())).to_dict()

            for idx, row in df.iterrows():
                sid = str(row[src_col]).strip()
                if sid in src_info and src_info[sid]["type"] == "child_id":
                    parent_id = id_map.get(sid, "0000000000")
                    parent_ekorgs = ekorg_map.get(parent_id, set())
                    ekorg_val = str(row[ekorg_col]).strip()
                    if ekorg_val and ekorg_val not in parent_ekorgs:
                        update_cell(df, idx, action_col, "I", modified_cells, sheet_name)
                        update_cell(df, idx, src_col, parent_id, modified_cells, sheet_name)

            sheets[sheet_name] = df
    else:
        print("ℹ️ LFM1 - Purchasing Org Data not found (skipped).")

    # ------------------------------------------------------------
    # Step 8: WYT3 - Partner Function (Supplier)
    # ------------------------------------------------------------
    sheet_name = "WYT3 - Partner Function (Suppli"
    hidden_cols = ["ERNAM", "ERDAT", "LIFN2", "LIFNR"]

    if sheet_name in sheets:
        print(f"🛠 Updating {sheet_name}...")
        df = clean_headers(sheets[sheet_name])
        try:
            src_col = find_column(df, "Source_ID")
            ekorg_col = find_column(df, "EKORG")
        except KeyError as e:
            print(f"⚠️ {e}")
            src_col = ekorg_col = None

        if src_col and ekorg_col:
            action_col = ensure_column(df, "_ACTION_CODE")
            ekorg_map = df.groupby(src_col)[ekorg_col].apply(lambda x: set(x.dropna().astype(str).str.strip())).to_dict()

            parvw_col = next((c for c in df.columns if c.strip().lower() == "parvw"), None)
            defpa_col = ensure_column(df, "DEFPA")

            for idx, row in df.iterrows():
                sid = str(row[src_col]).strip()
                if sid in src_info and src_info[sid]["type"] == "child_id":
                    parent_id = id_map.get(sid, "0000000000")
                    parent_ekorgs = ekorg_map.get(parent_id, set())
                    ekorg_val = str(row[ekorg_col]).strip()
                    if ekorg_val and ekorg_val not in parent_ekorgs:
                        update_cell(df, idx, action_col, "I", modified_cells, sheet_name)
                        update_cell(df, idx, src_col, parent_id, modified_cells, sheet_name)

                if parvw_col and str(row[parvw_col]).strip().upper() == "LF":
                    update_cell(df, idx, defpa_col, "X", modified_cells, sheet_name)

            sheets[sheet_name] = df
    else:
        print("ℹ️ WYT3 - Partner Function (Supplier) not found (skipped).")

    # ------------------------------------------------------------
    # Step 9: Save & highlight (unchanged)
    # ------------------------------------------------------------
    print("💾 Saving results and applying highlights...")

    input_file.seek(0)
    first_rows = pd.read_excel(input_file, sheet_name=None, nrows=1, header=None, dtype=str)

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        for name, df in sheets.items():
            first_row = first_rows.get(name)
            if first_row is not None:
                first_row.to_excel(writer, index=False, sheet_name=name, header=False)
                df.to_excel(writer, index=False, sheet_name=name, startrow=1, header=True)
            else:
                df.to_excel(writer, index=False, sheet_name=name, startrow=1, header=True)

    wb = load_workbook(output_file)
    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    changed_cols = {}
    for sheet_name, r_idx, col_name in modified_cells:
        ws = wb[sheet_name]
        headers = {cell.value: i + 1 for i, cell in enumerate(ws[2]) if cell.value}
        if col_name in headers:
            cell = ws.cell(row=r_idx + 3, column=headers[col_name])
            cell.fill = fill
            changed_cols.setdefault(sheet_name, set()).add(col_name)

    for ws in wb.worksheets:
        if ws.title not in required_sheets + optional_sheets or ws.title == "BUT100 - Role":
            ws.sheet_state = "hidden"
            
    # Hide columns except Source_ID + changed ones
    # === Hide columns based on specific rules ===
    for ws in wb.worksheets:
        sheet_name = ws.title
        header_cells = ws[2]

        # Rule 1️⃣: For "BUT000 - General" and "ADRC - Address"
        if sheet_name in ["BUT000 - General", "ADRC - Address"]:
            visible = set(changed_cols.get(sheet_name, set()))

            # Always include "Source_ID" column
            for cell in header_cells:
                if cell.value and "source" in str(cell.value).lower() and "id" in str(cell.value).lower():
                    visible.add(str(cell.value).strip())

            # Hide all columns not in visible set
            for cell in header_cells:
                if cell.value and str(cell.value).strip() not in visible:
                    ws.column_dimensions[cell.column_letter].hidden = True

        # Rule 2️⃣: For "WYT3 - Partner Function (Suppli)"
        elif sheet_name == "WYT3 - Partner Function (Suppli":
            for cell in header_cells:
                if str(cell.value).strip().upper() in ["ERNAM", "ERDAT", "LIFN2", "LIFNR"]:
                    ws.column_dimensions[cell.column_letter].hidden = True

        # Rule 3️⃣: For all other sheets (LFA1, LFB1, LFM1, etc.) — show all columns
        else:
            for cell in header_cells:
                if cell.value:
                    ws.column_dimensions[cell.column_letter].hidden = False
            
        # ------------------------------------------------------------
    # Step 11: Final styling for row 1 and row 2
    # ------------------------------------------------------------
    border_style = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )
    fill_style = PatternFill(start_color="DBD5BF", end_color="DBD5BF", fill_type="solid")

    for ws in wb.worksheets:
        max_col = ws.max_column
        # Apply to row 1 and 2
        for row_idx in [1, 2]:
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = border_style
                cell.fill = fill_style

    wb.save(output_file)
    return output_file


# ============================================================
#                 STREAMLIT UI
# ============================================================

st.set_page_config(page_title="Excel Vendor Processor", page_icon="📊", layout="centered")

st.title("📊 Common Supplier Upload File")
st.write("Input the pre-file from Fiori and generate the Upload file.")

st.subheader("🔗 Define Parent–Child Relationships")

# Step 1: Ask how many mappings
num_pairs = st.number_input("How many parent–child pairs?", min_value=1, max_value=50, value=1, step=1)

# Step 2: Let user fill in each pair
pairs = []
for i in range(int(num_pairs)):
    cols = st.columns(2)
    parent_id = cols[0].text_input(f"Parent ID #{i+1} (10 digits inc. leading zero)" )
    child_id = cols[1].text_input(f"Child ID #{i+1} (10 digits inc. leading zero)")
    if parent_id and child_id:
        pairs.append((parent_id.strip(), child_id.strip()))

st.markdown("---")
uploaded = st.file_uploader("📂 Upload Pre file (.xlsx)", type=["xlsx"])
progress_placeholder = st.empty()

if uploaded:
    try:
        # Read the file first to check Source_IDs
        preview_file = BytesIO(uploaded.read())
        sheets_preview = pd.read_excel(preview_file, sheet_name=None, header=1, dtype=str)
        but000 = clean_headers(sheets_preview["BUT000 - General"])
        source_ids_in_file = set(but000[find_column(but000, "Source_ID")].dropna().astype(str).str.strip())

        errors = []

        # --- Check 1: Format (10 digits)
        for i, (parent_id, child_id) in enumerate(pairs, 1):
            if not (len(parent_id) == 10 and parent_id.isdigit()):
                errors.append(f"❌ Pair #{i}: Parent ID '{parent_id}' must be exactly 10 digits.")
            if not (len(child_id) == 10 and child_id.isdigit()):
                errors.append(f"❌ Pair #{i}: Child ID '{child_id}' must be exactly 10 digits.")

        # --- Check 2: Parent 4th char = '3' (DELETED BECAUSE PARENT ID IS NOT NECESSARILY MUST BE ATLAS CODE)
        #for i, (parent_id, _) in enumerate(pairs, 1):
        #    if len(parent_id) == 10 and parent_id[3] != "3":
        #        errors.append(f"❌ Pair #{i}: Parent ID '{parent_id}' must have '3' as the 4th character.")

        # --- Check 3: Existence in file
        for i, (parent_id, child_id) in enumerate(pairs, 1):
            if parent_id not in source_ids_in_file:
                errors.append(f"❌ Pair #{i}: Parent ID '{parent_id}' not found in Source_ID column.")
            if child_id not in source_ids_in_file:
                errors.append(f"❌ Pair #{i}: Child ID '{child_id}' not found in Source_ID column.")

        if errors:
            st.error("Please fix the following issues before proceeding:")
            for e in errors:
                st.write(e)
        else:
            st.success("✅ All IDs validated successfully!")

            if st.button("Generate Upload File"):
                try:
                    with st.spinner("Processing... Please wait."):
                        processed_file = process_excel(preview_file.getvalue(), pairs)
                    st.success("✅ Done! Click below to download your file:")
                    st.download_button(
                        label="⬇️ Download Processed File",
                        data=processed_file,
                        file_name="UploadFile.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                except Exception as e:
                    st.error(f"❌ Error during processing: {e}")

    except Exception as e:
        st.error(f"❌ Failed to read Excel file or Source_ID column: {e}")
else:
    st.info("ℹ️ Please provide both the parent–child pairs and upload a file to proceed.")



